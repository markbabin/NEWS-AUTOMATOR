"""
Excel output module.
Writes detected topic segments to an Excel file with proper formatting.
Supports per-topic extra fields (e.g. Kolo, Tekma for PLT).
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_HEADERS = ["Datum", "Kanal", "Oddaja", "Tema"]
TAIL_HEADERS = ["Začetek", "Konec"]

COL_WIDTHS = {
    "Datum": 14,
    "Kanal": 12,
    "Oddaja": 22,
    "Tema": 18,
    "Začetek": 12,
    "Konec": 12,
    "Kolo": 8,
    "Tekma": 14,
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def get_all_extra_field_names(topics: list[dict]) -> list[str]:
    """Collect all unique extra field names across all topics, in order of first appearance."""
    seen = []
    for t in topics:
        for ef in t.get("extra_fields", []):
            if ef["name"] not in seen:
                seen.append(ef["name"])
    return seen


def get_extra_fields_for_topic(topic_name: str, topics: list[dict]) -> list[str]:
    """Get extra field names for a specific topic."""
    for t in topics:
        if t["name"] == topic_name:
            return [ef["name"] for ef in t.get("extra_fields", [])]
    return []


def _write_headers(ws, all_headers: list[str]) -> None:
    ws.append(all_headers)
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 14)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def get_or_create_workbook(
    output_path: Path,
    topics: list[dict],
) -> tuple[openpyxl.Workbook, any, list[str]]:
    """
    Load existing workbook or create a new one.
    Adds any missing extra-field columns to an existing file.
    Returns (workbook, worksheet, all_headers).
    """
    extra_field_names = get_all_extra_field_names(topics)
    all_headers = BASE_HEADERS + extra_field_names + TAIL_HEADERS

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for name in extra_field_names:
            if name not in existing:
                col_idx = ws.max_column + 1
                cell = ws.cell(1, col_idx)
                cell.value = name
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(name, 14)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Segmenti"
        _write_headers(ws, all_headers)

    return wb, ws, all_headers


def append_rows(
    output_path: Path,
    date: str,
    channel: str,
    show: str,
    topic_segments: dict[str, list[dict]],
    topics: list[dict],
) -> int:
    """
    Append rows for all detected segments to the Excel file.
    Returns the number of rows added.
    """
    wb, ws, all_headers = get_or_create_workbook(output_path, topics)
    extra_field_names = get_all_extra_field_names(topics)

    # Collect all rows first so we can sort by start timecode
    pending_rows = []
    for topic_name, segments in topic_segments.items():
        topic_extra_fields = get_extra_fields_for_topic(topic_name, topics)
        for seg in segments:
            # Use Oznaka as the Tema label if present (gender-tagged topics)
            tema = seg.get("Oznaka") or topic_name
            row_data = [date, channel, show, tema]
            for col_name in extra_field_names:
                if col_name in topic_extra_fields:
                    row_data.append(seg.get(col_name, ""))
                else:
                    row_data.append("")
            row_data += [seg["start"], seg["end"]]
            pending_rows.append(row_data)

    # Sort by start timecode (second-to-last column)
    pending_rows.sort(key=lambda r: r[-2])

    rows_added = 0
    current_row = ws.max_row + 1

    for row_data in pending_rows:
        ws.append(row_data)

        if current_row % 2 == 0:
            for col_idx in range(1, len(all_headers) + 1):
                ws.cell(row=current_row, column=col_idx).fill = ALT_ROW_FILL

        for col_idx in range(1, len(all_headers) + 1):
            ws.cell(row=current_row, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        current_row += 1
        rows_added += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return rows_added


def is_already_processed(output_path: Path, date: str, channel: str, show: str) -> bool:
    """Check if this date/channel/show combo already has rows in the Excel file."""
    if not output_path.exists():
        return False
    wb = openpyxl.load_workbook(output_path, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == date and row[1] == channel and row[2] == show:
            wb.close()
            return True
    wb.close()
    return False
